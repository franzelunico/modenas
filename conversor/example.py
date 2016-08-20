from openpyxl import load_workbook
from conversor import Conversor
# wb = Workbook()
wb = load_workbook('sample.xlsx')
distribucion = wb.get_sheet_by_name('distribucion')
encabezado_a = ["", ""]
encabezado_a += ["Billetes", "", "", "", ""]
encabezado_a += ["Monedas", "", ""]
encabezado_a += ["Centavos", "", ""]
encabezado_a += ["Desperdicio"]
distribucion.append(encabezado_a)  # Rows can also be appended
encabezado = ["Nombre", "Liquido Pagable"]
encabezado += [10, 20, 50, 100, 200]  # Billetes
encabezado += [1, 2, 5]  # Monedas
encabezado += [0.10, 0.20, 0.50]  # centavos
encabezado += [0]
distribucion.append(encabezado)  # Rows can also be appended

planilla = wb.get_sheet_by_name('planilla')

flag = True
fila = 2
nombre = "default value"
liquido_pagable = ""
res = ['Total', 'Distribucion', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

while(flag):
    nombre = planilla.cell(row=fila, column=1).value
    liquido_pagable = planilla.cell(row=fila, column=2).value
    if nombre is None or liquido_pagable is None:
        flag = False
    else:
        obrero = Conversor(nombre, liquido_pagable)
        datos = obrero.getDatos()
        distribucion.append(datos)  # Rows can also be appended
        for i in range(2, len(datos)):
            newvalue = res[i]+datos[i]
            res[i] = newvalue
        fila += 1
distribucion.append(res)  # Rows can also be appended
print fila

wb.save("sample.xlsx")  # Save the file
