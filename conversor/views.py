from django.shortcuts import render
from django.conf import settings
from django.http import HttpResponse
from .models import Document, UploadForm
from openpyxl import load_workbook
from conversor import Conversor
import os
# Create your views here.


def index_page(request):
    if request.method == 'POST':
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            up_docfile = request.FILES['docfile'].__str__()
            newdoc = Document(docfile=request.FILES['docfile'])
            newdoc.save(form)
            path = settings.MEDIA_ROOT
            path += '/documents/' + up_docfile
            # agregar datos
            wb = load_workbook(path)
            distribucion = wb.get_sheet_by_name('Conversor')
            encabezado_a = ["", ""]
            encabezado_a += ["Billetes", "", "", "", ""]
            encabezado_a += ["Monedas", "", ""]
            encabezado_a += ["Centavos", "", ""]
            encabezado_a += ["Desperdicio"]
            distribucion.append(encabezado_a)  # Rows can also be appended
            encabezado = ["Nombre", "Liquido Pagable"]
            encabezado += [10, 20, 50, 100, 200]  # Billetes
            encabezado += [1, 2, 5]  # Monedas
            encabezado += [0.10, 0.20, 0.50]  # centavos
            encabezado += [0]
            distribucion.append(encabezado)  # Rows can also be appended

            planilla = wb.get_sheet_by_name('Planilla')

            flag = True
            fila = 2
            nombre = "default value"
            liquido_pagable = ""
            res = ['Total', 'Distribucion', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

            while(flag):
                nombre = planilla.cell(row=fila, column=1).value
                liquido_pagable = planilla.cell(row=fila, column=2).value
                if nombre is None or liquido_pagable is None:
                    flag = False
                else:
                    obrero = Conversor(nombre, liquido_pagable)
                    datos = obrero.getDatos()
                    distribucion.append(datos)  # Rows can also be appended
                    for i in range(2, len(datos)):
                        newvalue = res[i]+datos[i]
                        res[i] = newvalue
                    fila += 1
            fila += 1
            distribucion.append(res)  # Rows can also be appended
            print fila

            wb.save(path)  # Save the file
            # estructurar el archivo
            my_data = ""
            with open(path, 'r') as f:
                my_data = f.read()
            response = HttpResponse(my_data,
                                    content_type='application/vnd.ms-excel')
            res_file = 'attachment; ' + 'filename="' + up_docfile + '"'
            response['Content-Disposition'] = res_file
            os.remove(path)
            return response
    else:
        form = UploadForm()
    return render(request, 'index.html', {'form': form})


def ejemplo(request):
    path = settings.MEDIA_ROOT
    path += '/ejemplo/sample.xlsx'
    my_data = ""
    with open(path, 'r') as f:
        my_data = f.read()
    response = HttpResponse(my_data,
                            content_type='application/vnd.ms-excel')
    res_file = 'attachment; ' + 'filename="' + 'sample.xlsx' + '"'
    response['Content-Disposition'] = res_file
    print "Descargar"
    return response
